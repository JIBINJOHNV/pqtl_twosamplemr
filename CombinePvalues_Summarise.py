import glob,os
import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill



def save_dataframe_to_excel_with_highlighting(dataframe, output_filename):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    
    # Create fill patterns for yellow and another color
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    other_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green color
    
    # Write column headers
    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = col_name
    
    # Write data to the worksheet
    for row_idx, row in enumerate(dataframe.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value
    
    # Apply highlighting based on the condition
    for col_idx, col_name in enumerate(dataframe.columns[1:], start=2):
        if col_name.endswith('LwestPvalue'):
            for row_idx, value in enumerate(dataframe[col_name], start=2):
                if pd.notna(value) and value < 0.000099:
                    worksheet.cell(row=row_idx, column=col_idx).fill = other_fill  # Use the other fill color
        elif col_name.endswith(('Pvalue', 'metap')):
            for row_idx, value in enumerate(dataframe[col_name], start=2):
                if pd.notna(value) and value < 0.000099:
                    worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
    # Save the Excel file
    workbook.save(output_filename)



def merge_files(files_withdict,prefix1):
    master_df=pd.DataFrame(columns=["Gene_Symbol"])
    for file_1 in files_withdict.keys():
        df=pd.read_csv(file_1)
        df=df[["Gene_Symbol"]].drop_duplicates()
        master_df=pd.merge(master_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    
    for file in files_withdict.keys():
        df=pd.read_csv(file)
        df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "MRIVWtest"  in x]]
        df2.columns=[x.replace(":BritishMR.IVWDelta","") for x in df2.columns ]
        df2.columns=[x.replace("_MRIVWtest","") for x in df2.columns ]
        df2.columns=["Gene_Symbol"]+[files_withdict[file]+x for x in df2.columns if "Gene_Symbol" not in x]
        column_name1=files_withdict[file]+"LwestPvalue"
        df2[column_name1]=df2.iloc[:,1:].min(axis=1)
        df2=df2.sort_values(column_name1).drop_duplicates(subset='Gene_Symbol', keep='first')
        master_df=pd.merge(master_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)
    
    min_pvaluecolumns=[x for x in master_df.columns if "_LwestPvalue" in x ]
    master_df[f"{prefix1}_Significant_count"]=master_df[min_pvaluecolumns].apply(lambda row: (row < 0.000099).sum(), axis=1)
    pvalue_columns=sorted([ x for x in master_df.columns if 'Gene_Symbol' not in x and "Significant_count" not in x])
    master_df=master_df[['Gene_Symbol']+pvalue_columns+[f"{prefix1}_Significant_count"]]
    filtered_df = master_df[(master_df[pvalue_columns] < 0.000099).any(axis=1)]
    out_exceellName=f"{prefix1}_Biogen_Decode_MR.xlsx"
    save_dataframe_to_excel_with_highlighting(filtered_df,out_exceellName)
    return master_df



##cis
cis_files_dict={'Biogen_Decode_pQTL_Depression_iPSYCH_2023_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"Depression_Cis_", 
                'Biogen_Decode_pQTL_PGC3_SCZ_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"SCZ_Cis_", 
                'Biogen_Decode_pQTL_BIP_PGC3_noukb_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"BIP_Cis_"}

cis_df=merge_files(cis_files_dict,"Cis")


##Trans with MHC
TransWithmhc_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_TransWithMHC_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransWithMHC_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransWithMHC_"}
TransWithmhc_files_df=merge_files(TransWithmhc_files_dict,"TransWithMHC")


##TransNoMHC
TransWithoutmhc_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_TransNoMHC_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransNoMHC_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransNoMHC_"}

TransWithoutmhc_files_df=merge_files(TransWithoutmhc_files_dict,"TransWithoutMHC")


##TransNoMHC
TransWithoutmhcUnique_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_ransNoMHCUnique_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransNoMHCUnique_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransNoMHCUnique_"}
TransWithoutmhcUnique_files_df=merge_files(TransWithoutmhcUnique_files_dict,"TransWithoutMHCUnique")


master_df=pd.DataFrame(columns=["Gene_Symbol"])

for file_1 in glob.glob("*MetapAnalysis.csv"):
    df=pd.read_csv(file_1)
    df=df[["Gene_Symbol"]].drop_duplicates()
    master_df=pd.merge(master_df,df,on="Gene_Symbol",how="outer").drop_duplicates()

for df in [cis_df,TransWithmhc_files_df,TransWithoutmhc_files_df,TransWithoutmhcUnique_files_df]:
    master_df=pd.merge(master_df,df,on="Gene_Symbol",how="outer").drop_duplicates()

filtered_df = master_df[(master_df[[ x for x in master_df.columns if 'LwestPvalue' in x]] < 0.000099).any(axis=1)]

save_dataframe_to_excel_with_highlighting(master_df, "Biogen_Decode_SCZ_BIP_Depression_BritishMR_Pipeline_AllProteins.xls")
save_dataframe_to_excel_with_highlighting(filtered_df, "Biogen_Decode_SCZ_BIP_Depression_BritishMR_Pipeline_Pvzlue_0.000099.xls")

filtered_df.to_csv("Biogen_Decode_SCZ_BIP_Depression_BritishMR_Pipeline_Pvzlue_0.000099.csv",index=None)



