import glob,os
import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill



def save_dataframe_to_excel_with_highlighting(dataframe, output_filename):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    
    # Create fill patterns for yellow and another color
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    other_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green color
    
    # Write column headers
    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = col_name
    
    # Write data to the worksheet
    for row_idx, row in enumerate(dataframe.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value
    
    # Apply highlighting based on the condition
    for col_idx, col_name in enumerate(dataframe.columns[1:], start=2):
        if col_name.endswith('LwestPvalue'):
            for row_idx, value in enumerate(dataframe[col_name], start=2):
                if pd.notna(value) and value < 0.000099:
                    worksheet.cell(row=row_idx, column=col_idx).fill = other_fill  # Use the other fill color
        elif col_name.endswith(('Pvalue', 'metap')):
            for row_idx, value in enumerate(dataframe[col_name], start=2):
                if pd.notna(value) and value < 0.000099:
                    worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
    # Save the Excel file
    workbook.save(output_filename)


cis_files=glob.glob("*CisExposure_CompleteMR*")
cis_files_dict={'Biogen_Decode_pQTL_Depression_iPSYCH_2023_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"Depression_Cis_", 
                'Biogen_Decode_pQTL_PGC3_SCZ_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"SCZ_Cis_", 
                'Biogen_Decode_pQTL_BIP_PGC3_noukb_CisExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv':"BIP_Cis_"}

cis_df=pd.DataFrame(columns=["Gene_Symbol"])

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df=df[["Gene_Symbol"]].drop_duplicates()
    cis_df=pd.merge(cis_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "MRIVWtest"  in x]]
    df2.columns=[x.replace(":BritishMR.IVWDelta","") for x in df2.columns ]
    df2.columns=[x.replace("_MRIVWtest","") for x in df2.columns ]
    df2.columns=["Gene_Symbol"]+[cis_files_dict[cis_file]+x for x in df2.columns if "Gene_Symbol" not in x]
    df2[cis_files_dict[cis_file]+"LwestPvalue"]=df2.iloc[:,1:].min(axis=1)
    df2=df2.sort_values(cis_files_dict[cis_file]+"LwestPvalue").drop_duplicates(subset='Gene_Symbol', keep='first')
    cis_df=pd.merge(cis_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)

min_pvaluecolumns=[x for x in cis_df.columns if "_LwestPvalue" in x ]
cis_df["Cis_Significant_count"]=cis_df[min_pvaluecolumns].apply(lambda row: (row < 0.000099).sum(), axis=1)
#cis_df.drop(min_pvaluecolumns,axis=1,inplace=True)

filtered_df = cis_df[(cis_df.iloc[:, 1:-1] < 0.000099).any(axis=1)]

# Call the function to save the DataFrame to Excel with highlighting
save_dataframe_to_excel_with_highlighting(filtered_df, 'CisExposure_CompleteMR.xlsx')



##Trans with MHV
cis_files=glob.glob("*TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv")
cis_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_TransWithMHC_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransWithMHC_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposure_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransWithMHC_"}

cis_df=pd.DataFrame(columns=["Gene_Symbol"])

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df=df[["Gene_Symbol"]].drop_duplicates()
    cis_df=pd.merge(cis_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    
for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "MRIVWtest"  in x]]
    df2.columns=[x.replace(":BritishMR.IVWDelta","") for x in df2.columns ]
    df2.columns=[x.replace("_MRIVWtest","") for x in df2.columns ]
    df2.columns=["Gene_Symbol"]+[cis_files_dict[cis_file]+x for x in df2.columns if "Gene_Symbol" not in x]
    df2[cis_files_dict[cis_file]+"LwestPvalue"]=df2.iloc[:,1:].min(axis=1)
    df2=df2.sort_values(cis_files_dict[cis_file]+"LwestPvalue").drop_duplicates(subset='Gene_Symbol', keep='first')
    cis_df=pd.merge(cis_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)

min_pvaluecolumns=[x for x in cis_df.columns if "_LwestPvalue" in x ]
cis_df["WithMHC_Significant_count"]=cis_df[min_pvaluecolumns].apply(lambda row: (row < 0.000099).sum(), axis=1)
#cis_df.drop(min_pvaluecolumns,axis=1,inplace=True)

filtered_df = cis_df[(cis_df.iloc[:, 1:-1] < 0.000099).any(axis=1)]

# Call the function to save the DataFrame to Excel with highlighting
save_dataframe_to_excel_with_highlighting(filtered_df, 'WithMHC__Exposure_CompleteMR.xlsx')



##TransNoMHC
cis_files=glob.glob("*TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv")
cis_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_TransNoMHC_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransNoMHC_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposureNoMHC_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransNoMHC_"}

cis_df=pd.DataFrame(columns=["Gene_Symbol"])

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df=df[["Gene_Symbol"]].drop_duplicates()
    cis_df=pd.merge(cis_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "MRIVWtest"  in x]]
    df2.columns=[x.replace(":BritishMR.IVWDelta","") for x in df2.columns ]
    df2.columns=[x.replace("_MRIVWtest","") for x in df2.columns ]
    df2.columns=["Gene_Symbol"]+[cis_files_dict[cis_file]+x for x in df2.columns if "Gene_Symbol" not in x]
    df2[cis_files_dict[cis_file]+"LwestPvalue"]=df2.iloc[:,1:].min(axis=1)
    df2=df2.sort_values(cis_files_dict[cis_file]+"LwestPvalue").drop_duplicates(subset='Gene_Symbol', keep='first')
    cis_df=pd.merge(cis_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)

min_pvaluecolumns=[x for x in cis_df.columns if "_LwestPvalue" in x ]
cis_df["TransNoMHC_Significant_count"]=cis_df[min_pvaluecolumns].apply(lambda row: (row < 0.000099).sum(), axis=1)
#cis_df.drop(min_pvaluecolumns,axis=1,inplace=True)

filtered_df = cis_df[(cis_df.iloc[:, 1:-1] < 0.000099).any(axis=1)]

# Call the function to save the DataFrame to Excel with highlighting
save_dataframe_to_excel_with_highlighting(filtered_df, 'TransNoMHC_Exposure_CompleteMR.xlsx')




##TransNoMHC
cis_files=glob.glob("*TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv")
cis_files_dict = {
    'Biogen_Decode_pQTL_PGC3_SCZ_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "SCZ_ransNoMHCUnique_", 
    'Biogen_Decode_pQTL_Depression_iPSYCH_2023_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "Depression_TransNoMHCUnique_", 
    'Biogen_Decode_pQTL_BIP_PGC3_noukb_TransExposureNoMHCUnique_CompleteMR_AnalysisResults_WithselectedColumns_MetapAnalysis.csv': "BIP_TransNoMHCUnique_"}

cis_df=pd.DataFrame(columns=["Gene_Symbol"])

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df=df[["Gene_Symbol"]].drop_duplicates()
    cis_df=pd.merge(cis_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    

for cis_file in cis_files_dict.keys():
    df=pd.read_csv(cis_file)
    df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "MRIVWtest"  in x]]
    df2.columns=[x.replace(":BritishMR.IVWDelta","") for x in df2.columns ]
    df2.columns=[x.replace("_MRIVWtest","") for x in df2.columns ]
    df2.columns=["Gene_Symbol"]+[cis_files_dict[cis_file]+x for x in df2.columns if "Gene_Symbol" not in x]
    df2[cis_files_dict[cis_file]+"LwestPvalue"]=df2.iloc[:,1:].min(axis=1)
    df2=df2.sort_values(cis_files_dict[cis_file]+"LwestPvalue").drop_duplicates(subset='Gene_Symbol', keep='first')
    cis_df=pd.merge(cis_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)

min_pvaluecolumns=[x for x in cis_df.columns if "_LwestPvalue" in x ]
cis_df["ransNoMHCUnique__Significant_count"]=cis_df[min_pvaluecolumns].apply(lambda row: (row < 0.000099).sum(), axis=1)
#cis_df.drop(min_pvaluecolumns,axis=1,inplace=True)

filtered_df = cis_df[(cis_df.iloc[:, 1:-1] < 0.000099).any(axis=1)]

# Call the function to save the DataFrame to Excel with highlighting
save_dataframe_to_excel_with_highlighting(filtered_df, 'TransNoMHCUnique_Exposure_CompleteMR.xlsx')


