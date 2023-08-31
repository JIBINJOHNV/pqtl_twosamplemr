import glob,os,re
import pandas as pd
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font



def save_dataframe_to_excel_with_highlighting(dataframe, output_filename, gwas_names, pqtltypes,pqt_gwas_dict):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    
    # Create fill patterns for different colors
    pale_green_fill = PatternFill(fill_type='solid', fgColor='98FB98')  # Pale Green color
    khaki_fill = PatternFill(fill_type='solid', fgColor='F0E68C')  # Khaki color
    pink_fill = PatternFill(fill_type='solid', fgColor='FFC0CB')  # Pink color
    yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')  # Yellow color
    
    # Create font for white text color
    white_font = Font(color="FFFFFF")
    
    # Write column headers
    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=col_idx).value = col_name
    
    # Write data to the worksheet
    for row_idx, row in enumerate(dataframe.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value
    
    # Apply highlighting based on the conditions
    for col_idx, col_name in enumerate(dataframe.columns[1:], start=2):
        if col_name.endswith(('LwestPvalue', 'Pvalue', 'metap')):
            for row_idx, value in enumerate(dataframe[col_name], start=2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if pd.notna(value):
                    if value < 0.000099:
                        cell.fill = pale_green_fill
                    elif 0.000099 < value < 0.0009999:
                        cell.fill = khaki_fill
                    elif 0.0009999 < value < 0.01:
                        cell.fill = pink_fill
                    elif value > 0.01:
                        cell.font = white_font
    
    # Iterate through each pqtltype
    for pqtltype in pqtltypes:
        for gwas_name in gwas_names:
            beta_columns = [item for item in dataframe.columns if re.match(rf'{pqtltype}_.*{gwas_name}_Beta', item)]
            # Highlight beta columns based on the new condition
            for row_idx in range(2, len(dataframe) + 2):
                row = dataframe.iloc[row_idx - 2]
                if ((row[beta_columns].max() > 0) and (row[beta_columns].min() < 0)):
                    for col_name in beta_columns:
                        col_idx = dataframe.columns.get_loc(col_name) + 1
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        worksheet.cell(row=row_idx, column=dataframe.columns.get_loc(col_name) + 1).fill = yellow_fill
    
    # Highlight Gene_Symbol column based on LwestPvalue conditions
    gene_symbol_col_idx = dataframe.columns.get_loc('Gene_Symbol') + 1
    
    # Iterate through rows
    for row_idx in range(2, len(dataframe) + 2):
        lwestpvalues = []  # Store non-NaN LwestPvalue values for the current row
        
        # Iterate through columns
        for col_name in dataframe.columns:
            if 'LwestPvalue' in col_name:
                lwestpvalue = worksheet.cell(row=row_idx, column=dataframe.columns.get_loc(col_name) + 1).value
                if pd.notna(lwestpvalue):
                    lwestpvalues.append(lwestpvalue)
        
        # Apply fill conditions to the Gene_Symbol column cell
        cell = worksheet.cell(row=row_idx, column=gene_symbol_col_idx)
        if lwestpvalues:
            min_lwestpvalue = min(filter(lambda x: pd.notna(x), lwestpvalues))
            if min_lwestpvalue < 0.000099:
                cell.fill = pale_green_fill
            elif 0.000099 < min_lwestpvalue < 0.0009999:
                cell.fill = khaki_fill
            elif 0.0009999 < min_lwestpvalue < 0.01:
                cell.fill = pink_fill
    
    for key, value in pqt_gwas_dict.items():
        # Check if the key is present in the dataframe columns
        if key in dataframe.columns:
            col_idx = dataframe.columns.get_loc(key) + 1
            
            # Get the corresponding value column index
            value_col_idx = dataframe.columns.get_loc(value) + 1
            
            # Iterate through rows and highlight based on condition
            for row_idx in range(2, len(dataframe) + 2):
                key_cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                value_cell = worksheet.cell(row=row_idx, column=value_col_idx)
                if pd.notna(key_cell_value):
                    if key_cell_value < 0.000099:
                        value_cell.fill = pale_green_fill
                    elif 0.000099 < key_cell_value < 0.0009999:
                        value_cell.fill = khaki_fill
                    elif 0.0009999 < key_cell_value < 0.01:
                        value_cell.fill = pink_fill
    
    # Save the Excel file
    workbook.save(output_filename)



def pqtltype_merge(cis_files_dict,pqtltype,pvalue_cutof1,pvalue_cutof2,pqt_gwas_dict):
    cis_df=pd.DataFrame(columns=["Gene_Symbol"])
    
    for cis_file in cis_files_dict.keys():
        df=pd.read_csv(cis_file)
        df=df[["Gene_Symbol"]].drop_duplicates()
        cis_df=pd.merge(cis_df,df,on="Gene_Symbol",how="outer").drop_duplicates()
    
    for cis_file in cis_files_dict.keys():
        df=pd.read_csv(cis_file)
        df2=df[["Gene_Symbol"]+[x for x in df.columns if   "IVWDelta_Pvalue" in x or "metap_MRIVWtest"  in x  or "IVWDelta_SNPs"  in x or "IVWDelta_Beta" in x or "IVWDelta_nSNPs"  in x ]]
        df2.columns=[x.replace("_BritishMR.IVWDelta","") for x in df2.columns ]
        df2.columns=[x.replace("_MRIVWtest","_Pvalue") for x in df2.columns ]
        Lpvaluename=pqtltype+"_"+cis_files_dict[cis_file]+"_LwestPvalue"
        pvalue_columns=[x for x in df2.columns if   x.endswith("_Pvalue") ]
        df2[Lpvaluename]=df2[pvalue_columns].min(axis=1)
        df2=df2.sort_values(Lpvaluename).drop_duplicates(subset='Gene_Symbol', keep='first')
        df2[pqtltype+"_"+cis_files_dict[cis_file]+"_count"]=df2[pvalue_columns].apply(lambda row: (row < pvalue_cutof2).sum(), axis=1)
        cis_df=pd.merge(cis_df,df2,on="Gene_Symbol",how="outer").drop_duplicates() #.drop("LwestPvalue",axis=1)
    
    min_pvaluecolumns=[x for x in cis_df.columns if "_LwestPvalue" in x ]
    cis_df[pqtltype+"_TotallSignificant_count"]=cis_df[min_pvaluecolumns].apply(lambda row: (row < pvalue_cutof2).sum(), axis=1)
    filtered_df = cis_df[(cis_df[min_pvaluecolumns] < pvalue_cutof1).any(axis=1)]
    save_dataframe_to_excel_with_highlighting(cis_df, f'{pqtltype}Exposure_CompleteMR.xlsx', list(cis_files_dict.values()), [pqtltype],pqt_gwas_dict)
    return filtered_df
    
    

pqt_gwas_dict={'CIS_ASD_PGC_LwestPvalue': 'CIS_ASD_PGC_count', 'CIS_PGC3_SCZ_LwestPvalue': 'CIS_PGC3_SCZ_count', 'CIS_Depression_iPSYCH_2023_LwestPvalue': 'CIS_Depression_iPSYCH_2023_count', 'CIS_BIP_PGC3_noukb_LwestPvalue': 'CIS_BIP_PGC3_noukb_count', 'CIS_PGC_AN2_LwestPvalue': 'CIS_PGC_AN2_count', 'CIS_ADHD2022_iPSYCH_deCODE_LwestPvalue': 'CIS_ADHD2022_iPSYCH_deCODE_count',
               'TransExposureNoMHC_ASD_PGC_LwestPvalue': 'TransExposureNoMHC_ASD_PGC_count', 'TransExposureNoMHC_PGC3_SCZ_LwestPvalue': 'TransExposureNoMHC_PGC3_SCZ_count', 'TransExposureNoMHC_Depression_iPSYCH_2023_LwestPvalue': 'TransExposureNoMHC_Depression_iPSYCH_2023_count', 'TransExposureNoMHC_BIP_PGC3_noukb_LwestPvalue': 'TransExposureNoMHC_BIP_PGC3_noukb_count', 'TransExposureNoMHC_PGC_AN2_LwestPvalue': 'TransExposureNoMHC_PGC_AN2_count', 'TransExposureNoMHC_ADHD2022_iPSYCH_deCODE_LwestPvalue': 'TransExposureNoMHC_ADHD2022_iPSYCH_deCODE_count'}


pvalue_cutof1=0.05
pvalue_cutof2=0.01

all_pqtl_df=pd.DataFrame(columns=["Gene_Symbol"])

pqtltypes={"TransExposureNoMHCUnique":"TransExposureNoMHCUnique","CisExposure":"CIS","TransExposureNoMHC":"TransExposureNoMHC","TransExposure":"TransExposure"}

for pqtltype in pqtltypes.keys():
    cis_files_dict={f'Biogen_Decode_pQTL_ASD_PGC_{pqtltype}_British_IVDelt_MetapAnalysis.csv':'ASD_PGC', 
    f'Biogen_Decode_pQTL_PGC3_SCZ_{pqtltype}_British_IVDelt_MetapAnalysis.csv':"PGC3_SCZ", 
    f'Biogen_Decode_pQTL_Depression_iPSYCH_2023_{pqtltype}_British_IVDelt_MetapAnalysis.csv':"Depression_iPSYCH_2023", 
    f'Biogen_Decode_pQTL_BIP_PGC3_noukb_{pqtltype}_British_IVDelt_MetapAnalysis.csv':"BIP_PGC3_noukb", 
    f'Biogen_Decode_pQTL_PGC_AN2_{pqtltype}_British_IVDelt_MetapAnalysis.csv':"PGC_AN2", 
    f'Biogen_Decode_pQTL_PGC_ADHD2022_iPSYCH_deCODE_{pqtltype}_British_IVDelt_MetapAnalysis.csv':"ADHD2022_iPSYCH_deCODE"}
    
    pqtltype_2=pqtltypes[pqtltype]
    result_df=pqtltype_merge(cis_files_dict,pqtltype_2,pvalue_cutof1,pvalue_cutof2,pqt_gwas_dict)
    all_pqtl_df=pd.merge(all_pqtl_df,result_df,on="Gene_Symbol",how="outer").drop_duplicates()



all_pqtl_df_cis_nomhc=["Gene_Symbol"]+[x for x in all_pqtl_df.columns if x.startswith("CIS_") or x.startswith("TransExposureNoMHC_") ]
all_pqtl_df_cis_nomhc_df=all_pqtl_df[all_pqtl_df_cis_nomhc]


def order_function(unorderedcolumns):
    asdbeta=[x for x in unorderedcolumns if "ASD_PGC" in x]
    sczbeta=[x for x in unorderedcolumns if "PGC3_SCZ" in x]
    mddbeta=[x for x in unorderedcolumns if "Depression_iPSYCH_2023" in x]
    bipbeta=[x for x in unorderedcolumns if "BIP_PGC3_noukb" in x]
    eatingbeta=[x for x in unorderedcolumns if "PGC_AN2" in x]
    adhdeta=[x for x in unorderedcolumns if "ADHD2022_iPSYCH_deCODE" in x]
    beta_order=sczbeta+bipbeta+mddbeta+adhdeta+asdbeta+eatingbeta
    return(beta_order)

betacols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_Beta")])
lpvaluecols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_LwestPvalue")])
pvaluecols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_Pvalue")])
nsnpscols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_nSNPs")])
snpscols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_SNPs")])
countcols=order_function([x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("_count") and not x.endswith("TotallSignificant_count")   ])
totalcountcols=[x for x in all_pqtl_df_cis_nomhc_df.columns if x.endswith("TotallSignificant_count")]

ordered_cols=["Gene_Symbol"]+totalcountcols+countcols+lpvaluecols+pvaluecols+betacols+nsnpscols+snpscols
all_pqtl_df_cis_nomhc_df_ordered=all_pqtl_df_cis_nomhc_df[ordered_cols]


# Use str.contains to filter column names
ciscount_columns = all_pqtl_df_cis_nomhc_df_ordered.columns[all_pqtl_df_cis_nomhc_df_ordered.columns.str.contains(r"CIS_.*_count")]
all_pqtl_df_cis_nomhc_df_ordered['CIS_TotallSignificant_count']=all_pqtl_df_cis_nomhc_df_ordered[[x for x in ciscount_columns if "TotallSignificant_count" not in x]].sum(axis=1,skipna=True)

trans_columns = all_pqtl_df_cis_nomhc_df_ordered.columns[all_pqtl_df_cis_nomhc_df_ordered.columns.str.contains(r"TransExposureNoMHC_.*_count")]
all_pqtl_df_cis_nomhc_df_ordered['TransExposureNoMHC_TotallSignificant_count']=all_pqtl_df_cis_nomhc_df_ordered[[x for x in trans_columns if "TotallSignificant_count" not in x]].sum(axis=1,skipna=True)

save_dataframe_to_excel_with_highlighting(all_pqtl_df_cis_nomhc_df_ordered, f'CIS_TransExposureNoMHC_Exposure_CompleteMR.xlsx',list(cis_files_dict.values()),list(pqtltypes.values()),pqt_gwas_dict )






##Extract Nomhc AND Trnasexposure

all_pqtl_df_withmhc_unique=["Gene_Symbol"]+[x for x in all_pqtl_df.columns if x not in all_pqtl_df_cis_nomhc ]
all_pqtl_df_withmhc_unique_df=all_pqtl_df[all_pqtl_df_withmhc_unique]


betacols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_Beta")])
lpvaluecols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_LwestPvalue")])
pvaluecols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_Pvalue")])
nsnpscols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_nSNPs")])
snpscols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_SNPs")])
countcols=order_function([x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("_count") and not x.endswith("TotallSignificant_count")   ])
totalcountcols=[x for x in all_pqtl_df_withmhc_unique_df.columns if x.endswith("TotallSignificant_count")]

ordered_cols=["Gene_Symbol"]+totalcountcols+countcols+lpvaluecols+pvaluecols+betacols+nsnpscols+snpscols
all_pqtl_df_withmhc_unique_df_ordered=all_pqtl_df_withmhc_unique_df[ordered_cols]
save_dataframe_to_excel_with_highlighting(all_pqtl_df_withmhc_unique_df_ordered, f'TransExposureWithMHC_Unique_Exposure_CompleteMR.xlsx',list(cis_files_dict.values()),list(pqtltypes.values()),pqt_gwas_dict )



#Final_Results=pd.merge(all_pqtl_df_cis_nomhc_df_ordered,all_pqtl_df_withmhc_unique_df_ordered,on="Gene_Symbol")
#save_dataframe_to_excel_with_highlighting(Final_Results, f'TransExposureWithMHC_Unique_Exposure_CompleteMR.xlsx',list(cis_files_dict.values()),list(pqtltypes.values()),pqt_gwas_dict )

